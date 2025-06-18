from django.contrib.auth.models import User
from django.db.models import Q
from . import models
import os
import openpyxl
from datetime import datetime, date, time

ALLOWED_EXTENSIONS = ['xlsx']
MAX_FILE_SIZE = 10 * 1024 * 1024 # 10 MB

def check_user_by_username(username):
    return User.objects.filter(
        Q(username = username)).first()

def check_user_exists(username, email):
    """
    method to check whether user exists in database or not

    author - piyush gupta
    """
    return User.objects.filter(
        Q(username = username) | Q(email = email)).first()

def check_doctor_exists(doc_id):
    """
    method to check whether doctor exists in database or not

    author - piyush
    """
    return models.Doctor.objects.filter(doctor_id = doc_id).first()

def check_patient_exists(pat_id):
    """
    method to check whether patient exists in database or not

    author - piyush
    """
    return models.Patient.objects.filter(patient_id = pat_id).first()

def validate_file(file):
    ext = os.path.splitext(file.name)[1][1:].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return False, "File type not allowed for upload"
    
    if file.size > MAX_FILE_SIZE:
        return False, "File size is too large"
    
    return True, ""

def process_docavailability_bulkupload(uploaded_file):
    try:
        wb = openpyxl.load_workbook(filename=uploaded_file, read_only=True)
        sheet = wb.active

        required_headers = ['doctor_username', 'date', 'start_time', 'end_time']

        # validate headers
        first_row = next(sheet.iter_rows(min_row=1, max_row=1))
        actual_headers = [val.value for val in first_row]

        if actual_headers != required_headers:
            return None, f"Invalid headers. Expected: {required_headers}, Got: {actual_headers}"

        row_data = []
        row_errors = []
        seen_entries = set()

        for row_num, row in enumerate(sheet.iter_rows(min_row=2)):
            username_, date_, starttime_, endtime_ = row

            errors = []

            if is_blank(username_.value):
                errors.append("doctor_username cannot be empty")
            else:
                username_val = username_.value

            if not date_.value:
                errors.append("date cannot be empty")

            try:
                date_val = date_.value if isinstance(date_.value, date) else datetime.strptime(str(date_.value), '%d-%m-%Y')
            except ValueError:
                errors.append("date must be in DD-MM-YYYY format")

            if not is_correct_time_format(starttime_):
                errors.append("start_time must be in HH:MM format")
            else:
                start_val = starttime_.value if isinstance(starttime_.value, time) else datetime.strptime(str(starttime_.value), '%H:%M').time()

            if not is_correct_time_format(endtime_):
                errors.append("end_time must be in HH:MM format")
            else:
                end_val = endtime_.value if isinstance(endtime_.value, time) else datetime.strptime(str(endtime_.value), '%H:%M').time()

            if starttime_.value and endtime_.value:
                try:
                    start = start_val
                    end = end_val
                    if end <= start:
                        errors.append("end_time must be after start_time")
                except (ValueError, AttributeError):
                    errors.append("invalid time format for comparison")
            
            entry_key = (
                username_val,
                date_val,
                start_val,
                end_val)
            
            if entry_key in seen_entries:
                errors.append("duplicate entry for this doctor, date, start_time and end_time")
            else:
                seen_entries.add(entry_key)

            if errors:
                row_errors.append({
                    'row_number': row_num,
                    'errors': errors
                })
            else:
                row_data.append({
                    'doc_username': username_val,
                    'date': date_val,
                    'start_time' : start_val,
                    'end_time' : end_val}) 
        wb.close()

        if row_errors:
            return None, row_errors
        return row_data, None
    except Exception as e:
        if 'wb' in locals():
            wb.close()
        return None, f"Processing error: {str(e)}"
    
def is_correct_time_format(cell):
    if cell.data_type == 'd' and cell.number_format == 'h:mm':
        return True
    if isinstance(cell.value, time):
        return True
    if isinstance(cell.value, str) and ':' in cell.value:
        try:
            datetime.strptime(cell.value, '%H:%M')
            return True
        except ValueError:
            pass
    return False

def is_blank(value: str):
    if not value or not value.strip():
        return True
    return False

def create_individual_slot(data: dict):
    username = data.get('doc_username')
    doctor = check_doctor_by_username(username)

    if doctor is None:
        return False
    
    models.DoctorAvailability.objects.create(
        doctor = doctor,
        date = data.get('date'),
        start_time = data.get('start_time'),
        end_time = data.get('end_time'))
    return True
    
def check_doctor_by_username(username):
    return models.Doctor.objects.filter(user__username = username).first()

def check_patient_by_username(username):
    return models.Patient.objects.filter(user__username = username).first()